from openpyxl import load_workbook
from dataclasses import dataclass, asdict
from typing import List, Optional, Dict
from enum import Enum
import sys
import json
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from datetime import datetime, timezone

from pymongo import MongoClient
from pymongo.errors import PyMongoError, ServerSelectionTimeoutError


class Priority(Enum):
    HIGHEST = "Highest Priority"
    HIGH = "High Priority"
    NORMAL = "Normal"


@dataclass
class ProductionRecord:
    client_name: str
    item_description: str
    quantity: float
    priority: str

    def __str__(self) -> str:
        return (
            f"Client: {self.client_name:45} | "
            f"Item: {self.item_description:40} | "
            f"Qty: {self.quantity:8.0f} | "
            f"Priority: {self.priority}"
        )

    def to_dict(self) -> Dict:
        return asdict(self)


class ColorAnalyzer:

    @staticmethod
    def parse_rgb_color(rgb_string: Optional[str]) -> Optional[tuple]:
        if not rgb_string or rgb_string == "00000000":
            return None

        try:
            if isinstance(rgb_string, str) and len(rgb_string) == 8:
                rr = int(rgb_string[2:4], 16)
                gg = int(rgb_string[4:6], 16)
                bb = int(rgb_string[6:8], 16)
                return (rr, gg, bb)
        except (ValueError, IndexError):
            pass

        return None

    @staticmethod
    def detect_priority(rgb_string: Optional[str]) -> Priority:
        rgb = ColorAnalyzer.parse_rgb_color(rgb_string)

        if not rgb:
            return Priority.NORMAL

        r, g, b = rgb

        if b > 150 and r < 100:
            return Priority.HIGHEST

        if g > 150 and r < 100 and b < 100:
            return Priority.HIGH

        return Priority.NORMAL


class EmailSender:

    def __init__(self, sender_email: str, sender_password: str, recipient_email: str):
        self.sender_email = sender_email
        self.sender_password = sender_password
        self.recipient_email = recipient_email

    def send_report(self, records: List[ProductionRecord], file_path: str, worksheet_name: str,
                     attachment_path: Optional[str] = None) -> bool:
        try:
            msg = MIMEMultipart('mixed')
            msg['From'] = self.sender_email
            msg['To'] = self.recipient_email
            msg['Subject'] = f"Production Analysis Report - {datetime.now().strftime('%Y-%m-%d %H:%M')}"

            html_content = self._create_html_report(records, file_path, worksheet_name)

            html_part = MIMEText(html_content, 'html')
            msg.attach(html_part)

            if attachment_path and os.path.exists(attachment_path):
                with open(attachment_path, "rb") as f:
                    attachment = MIMEApplication(f.read(), _subtype="html")
                filename = os.path.basename(attachment_path)
                attachment.add_header('Content-Disposition', 'attachment', filename=filename)
                msg.attach(attachment)
                print(f"  📎 Attached '{filename}' to email")

            print("\n📧 Sending HTML email report...")
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
                server.login(self.sender_email, self.sender_password)
                server.send_message(msg)

            print(f"  ✅ Email sent successfully to {self.recipient_email}")
            return True

        except smtplib.SMTPAuthenticationError:
            print("❌ Email Error: Gmail credentials invalid")
            print("   Generate App Password: https://myaccount.google.com/apppasswords")
            return False
        except Exception as e:
            print(f"❌ Email Error: {str(e)}")
            return False

    def _create_html_report(self, records: List[ProductionRecord], file_path: str, worksheet_name: str) -> str:

        sorted_records = sorted(
            records,
            key=lambda r: (
                0 if r.priority == Priority.HIGHEST.value else
                1 if r.priority == Priority.HIGH.value else 2,
                r.client_name
            )
        )

        priority_counts = {}
        for record in records:
            priority_counts[record.priority] = priority_counts.get(record.priority, 0) + 1

        client_counts = {}
        for record in records:
            client_counts[record.client_name] = client_counts.get(record.client_name, 0) + 1

        total_qty = sum(record.quantity for record in records)

        table_rows = ""
        for record in sorted_records:
            priority_class = "priority-highest" if record.priority == Priority.HIGHEST.value else "priority-high" if record.priority == Priority.HIGH.value else "priority-normal"
            table_rows += f"""
            <tr class="{priority_class}">
                <td>{record.client_name}</td>
                <td>{record.item_description}</td>
                <td style="text-align: right;">{record.quantity:,.0f}</td>
                <td style="text-align: center;">{record.priority}</td>
            </tr>
            """

        top_clients = sorted(client_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        top_clients_html = ""
        for i, (client, count) in enumerate(top_clients, 1):
            top_clients_html += f"<tr><td>{i}</td><td>{client}</td><td>{count}</td></tr>"

        html = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Production Analysis Report</title>
            <style>
                * {{ margin: 0; padding: 0; box-sizing: border-box; }}
                body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f5f5f5; color: #333; line-height: 1.6; }}
                .container {{ max-width: 1200px; margin: 0 auto; background-color: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .header {{ background: linear-gradient(135deg, #2c3e50 0%, #34495e 100%); color: white; padding: 30px; border-radius: 8px; margin-bottom: 30px; text-align: center; }}
                .header h1 {{ font-size: 32px; margin-bottom: 10px; font-weight: 700; }}
                .header p {{ font-size: 14px; opacity: 0.9; }}
                .info-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px; margin-bottom: 30px; }}
                .info-box {{ background-color: #f8f9fa; border-left: 4px solid #3498db; padding: 15px; border-radius: 4px; }}
                .info-box label {{ font-weight: bold; color: #2c3e50; display: block; margin-bottom: 5px; }}
                .info-box value {{ color: #555; word-break: break-all; font-size: 13px; }}
                table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
                th {{ background-color: #2c3e50; color: white; padding: 15px; text-align: left; font-weight: 600; border: 1px solid #1a252f; }}
                td {{ padding: 12px 15px; border-bottom: 1px solid #ddd; }}
                tr:hover {{ background-color: #f9f9f9; }}
                .priority-highest {{ background-color: #fadbd8; border-left: 4px solid #e74c3c; }}
                .priority-highest td {{ font-weight: bold; color: #c0392b; }}
                .priority-high {{ background-color: #fdebd0; border-left: 4px solid #f39c12; }}
                .priority-high td {{ font-weight: 600; color: #d68910; }}
                .priority-normal {{ background-color: #fff; }}
                .priority-normal td {{ color: #555; }}
                .section-title {{ background-color: #ecf0f1; padding: 15px; margin-top: 20px; margin-bottom: 15px; border-left: 4px solid #3498db; font-size: 16px; font-weight: 600; color: #2c3e50; }}
                .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin-bottom: 30px; }}
                .stat-card {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 8px; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.1); }}
                .stat-card.highest {{ background: linear-gradient(135deg, #e74c3c 0%, #c0392b 100%); }}
                .stat-card.high {{ background: linear-gradient(135deg, #f39c12 0%, #d68910 100%); }}
                .stat-card.normal {{ background: linear-gradient(135deg, #27ae60 0%, #229954 100%); }}
                .stat-value {{ font-size: 28px; font-weight: bold; margin-bottom: 5px; }}
                .stat-label {{ font-size: 13px; opacity: 0.9; }}
                .footer {{ text-align: center; color: #999; font-size: 12px; margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; }}
                .generated-at {{ background-color: #fffacd; border-left: 4px solid #f1c40f; padding: 10px 15px; margin-bottom: 20px; border-radius: 4px; font-size: 13px; color: #856404; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>📊 Production Analysis Report</h1>
                    <p>Detailed Production Records & Analysis</p>
                </div>
                <div class="generated-at">
                    📅 Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
                </div>
                <div class="info-grid">
                    <div class="info-box"><label>Total Records:</label><value>{len(records)}</value></div>
                    <div class="info-box"><label>Total Quantity:</label><value>{total_qty:,.0f} units</value></div>
                    <div class="info-box"><label>Total Clients:</label><value>{len(client_counts)}</value></div>
                    <div class="info-box"><label>Worksheet:</label><value>{worksheet_name}</value></div>
                </div>
                <div class="section-title">📈 Priority Breakdown</div>
                <div class="stats-grid">
                    <div class="stat-card highest"><div class="stat-value">{priority_counts.get(Priority.HIGHEST.value, 0)}</div><div class="stat-label">Highest Priority</div></div>
                    <div class="stat-card high"><div class="stat-value">{priority_counts.get(Priority.HIGH.value, 0)}</div><div class="stat-label">High Priority</div></div>
                    <div class="stat-card normal"><div class="stat-value">{priority_counts.get(Priority.NORMAL.value, 0)}</div><div class="stat-label">Normal Priority</div></div>
                </div>
                <div class="section-title">📋 Production Records</div>
                <table>
                    <thead><tr><th>Client Name</th><th>Item Description</th><th style="text-align: right;">Quantity</th><th style="text-align: center;">Priority</th></tr></thead>
                    <tbody>{table_rows}</tbody>
                </table>
                <div class="section-title">🏆 Top 5 Clients by Order Count</div>
                <table>
                    <thead><tr><th style="width: 50px;">Rank</th><th>Client Name</th><th style="width: 100px; text-align: right;">Orders</th></tr></thead>
                    <tbody>{top_clients_html}</tbody>
                </table>
                <div class="footer">
                    <p>This report was automatically generated by Production Analysis System</p>
                    <p>Do not reply to this email. For queries, contact your administrator.</p>
                </div>
            </div>
        </body>
        </html>
        """

        return html


class MongoStorage:

    def __init__(self, mongo_uri: str, db_name: str = "production_db",
                 collection_name: str = "production_records"):
        self.mongo_uri = mongo_uri
        self.db_name = db_name
        self.collection_name = collection_name
        self.client = None
        self.db = None
        self.collection = None

    def connect(self) -> bool:
        try:
            self.client = MongoClient(self.mongo_uri, serverSelectionTimeoutMS=5000)
            self.client.admin.command("ping")
            self.db = self.client[self.db_name]
            self.collection = self.db[self.collection_name]
            print(f"  ✅ Connected to MongoDB -> db='{self.db_name}', collection='{self.collection_name}'")
            return True
        except ServerSelectionTimeoutError:
            print("❌ MongoDB Error: Could not connect to server (check URI / network / whitelist)")
            return False
        except PyMongoError as e:
            print(f"❌ MongoDB Error: {str(e)}")
            return False

    @staticmethod
    def build_json_payload(records: List[ProductionRecord], file_path: str, worksheet_name: str) -> Dict:
        generated_at = datetime.now(timezone.utc)

        payload = {
            "file_path": file_path,
            "worksheet_name": worksheet_name,
            "generated_at": generated_at.isoformat(),
            "total_records": len(records),
            "total_quantity": sum(r.quantity for r in records),
            "records": [r.to_dict() for r in records],
        }
        return payload

    def insert_records(self, records: List[ProductionRecord], file_path: str, worksheet_name: str,
                        as_single_document: bool = False) -> bool:
        if self.collection is None:
            print("❌ MongoDB Error: Not connected. Call connect() first.")
            return False

        try:
            generated_at = datetime.now(timezone.utc)

            if as_single_document:
                doc = self.build_json_payload(records, file_path, worksheet_name)
                result = self.collection.insert_one(doc)
                print(f"  ✅ Inserted 1 report document (id: {result.inserted_id})")
            else:
                docs = []
                for r in records:
                    doc = r.to_dict()
                    doc["file_path"] = file_path
                    doc["worksheet_name"] = worksheet_name
                    doc["generated_at"] = generated_at
                    docs.append(doc)

                if not docs:
                    print("⚠️  No records to insert.")
                    return True

                result = self.collection.insert_many(docs)
                print(f"  ✅ Inserted {len(result.inserted_ids)} records into MongoDB")

            return True
        except PyMongoError as e:
            print(f"❌ MongoDB Error: Failed to insert records - {str(e)}")
            return False

    def close(self) -> None:
        if self.client:
            self.client.close()


class ExcelProcessor:

    def __init__(self, file_path: str, worksheet_name: str = "Monthly Production Detail"):
        self.file_path = file_path
        self.worksheet_name = worksheet_name
        self.workbook = None
        self.worksheet = None
        self.client_headers: Dict[int, tuple] = {}
        self.records: List[ProductionRecord] = []

    def load_workbook(self) -> bool:
        try:
            self.workbook = load_workbook(self.file_path, data_only=True)
        except FileNotFoundError:
            print(f"❌ Error: File not found - '{self.file_path}'")
            return False
        except Exception as e:
            print(f"❌ Error: Invalid workbook - {str(e)}")
            return False

        return True

    def load_worksheet(self) -> bool:
        if not self.workbook:
            print("❌ Error: Workbook not loaded")
            return False

        try:
            self.worksheet = self.workbook[self.worksheet_name]
        except KeyError:
            available = ", ".join(self.workbook.sheetnames)
            print(f"❌ Error: Worksheet '{self.worksheet_name}' not found")
            print(f"   Available worksheets: {available}")
            return False

        return True

    def extract_client_headers(self) -> bool:
        try:
            client_header_row = 3
            client_start_column = 9

            for col in range(client_start_column, self.worksheet.max_column + 1):
                cell = self.worksheet.cell(row=client_header_row, column=col)

                if cell.value is None or str(cell.value).strip() == "":
                    continue

                rgb_color = None
                if cell.fill and hasattr(cell.fill, 'fgColor'):
                    rgb_color = cell.fill.fgColor.rgb

                priority = ColorAnalyzer.detect_priority(rgb_color)
                client_name = str(cell.value).strip()

                self.client_headers[col] = (client_name, priority)

            return True
        except Exception as e:
            print(f"❌ Error: Failed to extract client headers - {str(e)}")
            return False

    def extract_item_descriptions(self) -> Dict[int, str]:
        items = {}
        item_column = 3
        data_start_row = 4

        for row in range(data_start_row, self.worksheet.max_row + 1):
            cell = self.worksheet.cell(row=row, column=item_column)
            if cell.value is not None:
                items[row] = str(cell.value).strip()

        return items

    def process_quantities(self) -> bool:
        try:
            item_descriptions = self.extract_item_descriptions()
            data_start_row = 4

            for row in range(data_start_row, self.worksheet.max_row + 1):
                item_desc = item_descriptions.get(row, "Unknown Item")

                for col, (client_name, priority) in self.client_headers.items():
                    cell = self.worksheet.cell(row=row, column=col)
                    quantity = cell.value

                    if quantity is None or quantity == "":
                        continue

                    try:
                        qty_value = float(quantity) if isinstance(quantity, (int, float)) else float(str(quantity).replace("=", ""))
                    except (ValueError, AttributeError):
                        continue

                    if qty_value == 0:
                        continue

                    record = ProductionRecord(
                        client_name=client_name,
                        item_description=item_desc,
                        quantity=qty_value,
                        priority=priority.value
                    )
                    self.records.append(record)

            return True
        except Exception as e:
            print(f"❌ Error: Failed to process quantities - {str(e)}")
            return False

    def to_json(self, indent: int = 2) -> str:
        payload = MongoStorage.build_json_payload(self.records, self.file_path, self.worksheet_name)
        return json.dumps(payload, indent=indent, ensure_ascii=False)

    def save_json_file(self, output_path: str) -> bool:
        try:
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(self.to_json())
            print(f"  ✅ JSON saved to '{output_path}'")
            return True
        except Exception as e:
            print(f"❌ Error: Failed to save JSON file - {str(e)}")
            return False

    def save_manifest_html(self, output_path: str, template_path: Optional[str] = None) -> bool:
        try:
            if template_path is None:
                script_dir = os.path.dirname(os.path.abspath(__file__))
                template_path = os.path.join(script_dir, "manifest_template.html")

            if not os.path.exists(template_path):
                print(f"❌ Error: Template not found - '{template_path}'")
                print("   Place 'manifest_template.html' in the same folder as this script.")
                return False

            with open(template_path, "r", encoding="utf-8") as f:
                template = f.read()

            payload = MongoStorage.build_json_payload(self.records, self.file_path, self.worksheet_name)
            data_json = json.dumps(payload, ensure_ascii=False)

            html = template.replace("__REPORT_DATA_JSON__", data_json)

            with open(output_path, "w", encoding="utf-8") as f:
                f.write(html)

            print(f"  ✅ Manifest HTML saved to '{output_path}'")
            return True
        except Exception as e:
            print(f"❌ Error: Failed to save manifest HTML - {str(e)}")
            return False

    def print_report(self) -> None:
        if not self.records:
            print("⚠️  No records found with non-zero quantities.")
            return

        print("\n" + "="*160)
        print("PRODUCTION ANALYSIS REPORT")
        print("="*160)
        print(f"\nTotal Records: {len(self.records)}")
        print(f"Worksheet: {self.worksheet_name}")
        print(f"File: {self.file_path}")
        print("\n" + "-"*160)

        header = (
            f"{'Client Name':45} | "
            f"{'Item Description':40} | "
            f"{'Quantity':>8} | "
            f"{'Priority'}"
        )
        print(header)
        print("-"*160)

        sorted_records = sorted(
            self.records,
            key=lambda r: (
                0 if r.priority == Priority.HIGHEST.value else
                1 if r.priority == Priority.HIGH.value else 2,
                r.client_name
            )
        )

        for record in sorted_records:
            print(record)

        print("-"*160)
        self._print_summary_statistics()

    def _print_summary_statistics(self) -> None:
        print("\n📊 SUMMARY STATISTICS")
        print("-"*160)

        priority_counts = {}
        for record in self.records:
            priority_counts[record.priority] = priority_counts.get(record.priority, 0) + 1

        for priority in [Priority.HIGHEST.value, Priority.HIGH.value, Priority.NORMAL.value]:
            count = priority_counts.get(priority, 0)
            if count > 0:
                print(f"  {priority}: {count} records")

        client_counts = {}
        for record in self.records:
            client_counts[record.client_name] = client_counts.get(record.client_name, 0) + 1

        print(f"\n  Total Clients: {len(client_counts)}")
        print(f"  Top 5 Clients by Order Count:")
        for client, count in sorted(client_counts.items(), key=lambda x: x[1], reverse=True)[:5]:
            print(f"    - {client}: {count} items")

        total_qty = sum(record.quantity for record in self.records)
        print(f"\n  Total Quantity Ordered: {total_qty:,.0f} units")
        print("="*160 + "\n")

    def process(self, send_email: bool = True, sender_email: str = "", sender_password: str = "",
                recipient_email: str = "", save_to_mongo: bool = True, mongo_uri: str = "",
                mongo_db: str = "production_db", mongo_collection: str = "production_records",
                save_json_path: Optional[str] = None, manifest_html_path: Optional[str] = None) -> bool:
        print("🔄 Starting Production Analysis Processing...\n")

        print("Step 1: Loading workbook...")
        if not self.load_workbook():
            return False
        print("  ✅ Workbook loaded successfully")

        print("Step 2: Loading worksheet...")
        if not self.load_worksheet():
            return False
        print(f"  ✅ Worksheet '{self.worksheet_name}' loaded successfully")

        print("Step 3: Extracting client headers...")
        if not self.extract_client_headers():
            return False
        print(f"  ✅ Found {len(self.client_headers)} clients")

        print("Step 4: Processing quantities...")
        if not self.process_quantities():
            return False
        print(f"  ✅ Processed {len(self.records)} production records")

        print("Step 5: Generating report...\n")
        self.print_report()

        if save_json_path:
            print(f"Step 5b: Saving JSON to file '{save_json_path}'...")
            self.save_json_file(save_json_path)

        if manifest_html_path:
            print(f"Step 5c: Generating manifest HTML '{manifest_html_path}'...")
            self.save_manifest_html(manifest_html_path)

        if save_to_mongo and mongo_uri:
            print("Step 6: Saving records to MongoDB...")
            mongo_storage = MongoStorage(mongo_uri, mongo_db, mongo_collection)
            if mongo_storage.connect():
                mongo_storage.insert_records(
                    self.records, self.file_path, self.worksheet_name,
                    as_single_document=True
                )
                mongo_storage.close()

        if send_email and sender_email and sender_password and recipient_email:
            print("Step 7: Sending email report...")
            email_sender = EmailSender(sender_email, sender_password, recipient_email)
            email_sender.send_report(
                self.records, self.file_path, self.worksheet_name,
                attachment_path=manifest_html_path
            )

        return True


def main():
    file_path = r"C:\Users\Rahil khan\Downloads\monthly report.py\monthly production-analysis.xlsx"

    SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "rahilkhan784422@gmail.com")
    SENDER_PASSWORD = os.environ.get("SENDER_PASSWORD", "")
    RECIPIENT_EMAIL = os.environ.get("RECIPIENT_EMAIL", "rahilrrk18@gmail.com")

    MONGO_URI = os.environ.get("MONGO_URI", "mongodb://localhost:27017/")
    MONGO_DB = "production_db"
    MONGO_COLLECTION = "production_records"

    if not SENDER_PASSWORD:
        print("❌ Error: SENDER_PASSWORD environment variable is not set.")
        print("   Set it in PowerShell: $env:SENDER_PASSWORD='your-app-password'")
        sys.exit(1)

    processor = ExcelProcessor(
        file_path=file_path,
        worksheet_name="Monthly Production Detail"
    )

    if not processor.process(
        send_email=True,
        sender_email=SENDER_EMAIL,
        sender_password=SENDER_PASSWORD,
        recipient_email=RECIPIENT_EMAIL,
        save_to_mongo=True,
        mongo_uri=MONGO_URI,
        mongo_db=MONGO_DB,
        mongo_collection=MONGO_COLLECTION,
        save_json_path="production_records.json",
        manifest_html_path="production_manifest.html",
    ):
        sys.exit(1)

    sys.exit(0)


if __name__ == "__main__":
    main()
